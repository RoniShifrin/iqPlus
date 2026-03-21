"""Report export service — generates PDF and CSV student progress reports."""
import csv
import io
import logging
from datetime import datetime
from typing import List

logger = logging.getLogger(__name__)


async def generate_student_report(
    student_id: str, fmt: str = "pdf"
) -> tuple[bytes, str, str]:
    """
    Build a full student progress report.

    Returns: (content_bytes, content_type, filename)
    """
    from app.models import Course, Enrollment, EnrollmentStatusEnum
    from app.repositories import (
        LessonRecordRepository, ProgressMetricsRepository,
        AIAlertRepository, UserRepository,
    )

    student = await UserRepository.get_by_id(student_id)
    if not student:
        raise ValueError("Student not found")

    enrollments = await Enrollment.find(
        Enrollment.student_id == student_id,
        Enrollment.status == EnrollmentStatusEnum.ACTIVE,
    ).to_list()

    course_data: List[dict] = []
    for e in enrollments:
        course = await Course.get(e.course_id)
        if not course:
            continue
        records = await LessonRecordRepository.get_by_student_course(
            student_id, e.course_id
        )
        metrics = await ProgressMetricsRepository.get(student_id, e.course_id)
        alerts = await AIAlertRepository.list_by_student_course(
            student_id, e.course_id, limit=5
        )
        course_data.append(
            {"course": course, "records": records, "metrics": metrics, "alerts": alerts}
        )

    if fmt == "csv":
        return _build_csv(student, course_data)
    if fmt == "xlsx":
        return _build_xlsx(student, course_data)
    return _build_pdf(student, course_data)


# ── CSV ───────────────────────────────────────────────────────────────────────

def _build_csv(student, course_data: List[dict]) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["IQ PLUS — Student Progress Report"])
    w.writerow(["Student", student.full_name()])
    w.writerow(["Email", student.email])
    w.writerow(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    w.writerow([])

    for cd in course_data:
        course = cd["course"]
        w.writerow([f"Course: {course.name} ({course.code})"])

        m = cd["metrics"]
        if m:
            w.writerow(["Average Grade", f"{m.average_grade:.1f}%"])
            w.writerow(["Attendance Rate", f"{m.attendance_rate:.1f}%"])
            w.writerow(["Trend", m.trend_direction])
        w.writerow([])

        w.writerow(["Date", "Attendance", "Grade", "Teacher Feedback"])
        for r in cd["records"]:
            w.writerow([
                r.lesson_date.strftime("%Y-%m-%d"),
                r.attendance_status.value if hasattr(r.attendance_status, "value") else r.attendance_status,
                f"{r.grade_value:.1f}" if r.grade_value is not None else "",
                r.teacher_feedback or "",
            ])

        if cd["alerts"]:
            w.writerow([])
            w.writerow(["AI Alerts"])
            w.writerow(["Level", "Message", "Recommendation", "Date"])
            for a in cd["alerts"]:
                w.writerow([
                    a.alert_level.value if hasattr(a.alert_level, "value") else a.alert_level,
                    a.message,
                    a.recommendation,
                    a.created_at.strftime("%Y-%m-%d"),
                ])
        w.writerow([])

    fname = f"report_{student_id_from(student)}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return buf.getvalue().encode("utf-8"), "text/csv", fname


def student_id_from(student) -> str:
    return str(student.id)[:8]


# ── PDF ───────────────────────────────────────────────────────────────────────

def _build_pdf(student, course_data: List[dict]) -> tuple[bytes, str, str]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
        styles = getSampleStyleSheet()
        H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=18, spaceAfter=4)
        H2 = styles["Heading2"]
        H3 = styles["Heading3"]
        NL = styles["Normal"]

        elems = []
        elems.append(Paragraph("IQ PLUS — Student Progress Report", H1))
        elems.append(Paragraph(f"Student: {student.full_name()}", H2))
        elems.append(Paragraph(f"Email: {student.email}", NL))
        elems.append(
            Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", NL)
        )
        elems.append(Spacer(1, 0.5 * cm))

        BLUE = colors.HexColor("#2563eb")
        ORANGE = colors.HexColor("#f97316")
        LIGHT = colors.HexColor("#f8fafc")
        BORDER = colors.HexColor("#e2e8f0")

        for cd in course_data:
            course = cd["course"]
            elems.append(Paragraph(f"Course: {course.name} ({course.code})", H2))

            m = cd["metrics"]
            if m:
                t = Table(
                    [
                        ["Average Grade", f"{m.average_grade:.1f}%"],
                        ["Attendance Rate", f"{m.attendance_rate:.1f}%"],
                        ["Trend", m.trend_direction.capitalize()],
                    ],
                    colWidths=[5 * cm, 5 * cm],
                )
                t.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (0, -1), colors.lightblue),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ])
                )
                elems.append(t)
                elems.append(Spacer(1, 0.3 * cm))

            if cd["records"]:
                elems.append(Paragraph("Lesson Records", H3))
                rows = [["Date", "Attendance", "Grade", "Feedback"]]
                for r in cd["records"]:
                    att = r.attendance_status.value if hasattr(r.attendance_status, "value") else r.attendance_status
                    rows.append([
                        r.lesson_date.strftime("%Y-%m-%d"),
                        att,
                        f"{r.grade_value:.1f}" if r.grade_value is not None else "—",
                        (r.teacher_feedback or "")[:60],
                    ])
                t = Table(rows, colWidths=[3 * cm, 3 * cm, 2.5 * cm, 8.5 * cm])
                t.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
                        ("GRID", (0, 0), (-1, -1), 0.3, BORDER),
                    ])
                )
                elems.append(t)
                elems.append(Spacer(1, 0.3 * cm))

            if cd["alerts"]:
                elems.append(Paragraph("AI Alerts", H3))
                rows = [["Level", "Message", "Recommendation", "Date"]]
                for a in cd["alerts"]:
                    lvl = a.alert_level.value if hasattr(a.alert_level, "value") else a.alert_level
                    rows.append([
                        lvl,
                        a.message[:80],
                        a.recommendation[:80],
                        a.created_at.strftime("%Y-%m-%d"),
                    ])
                t = Table(rows, colWidths=[2.5 * cm, 5 * cm, 5 * cm, 3 * cm])
                t.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), ORANGE),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                    ])
                )
                elems.append(t)

            elems.append(Spacer(1, 0.5 * cm))

        doc.build(elems)
        buf.seek(0)
        fname = f"report_{student_id_from(student)}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        return buf.getvalue(), "application/pdf", fname

    except ImportError:
        logger.warning("reportlab not installed — falling back to CSV")
        return _build_csv(student, course_data)


# ── XLSX ──────────────────────────────────────────────────────────────────────

def _build_xlsx(student, course_data: List[dict]) -> tuple[bytes, str, str]:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        # Remove default sheet and build per-course sheets
        default_sheet = wb.active

        HEADER_FILL = PatternFill("solid", fgColor="2563EB")
        HEADER_FONT = Font(color="FFFFFF", bold=True)

        for cd in course_data:
            course = cd["course"]
            ws = wb.create_sheet(title=course.code[:31])

            # Title block
            ws.append([f"IQ PLUS — Student Progress Report"])
            ws.append([f"Student: {student.full_name()}"])
            ws.append([f"Email: {student.email}"])
            ws.append([f"Course: {course.name} ({course.code})"])
            ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
            ws.append([])

            m = cd["metrics"]
            if m:
                ws.append(["Average Grade", f"{m.average_grade:.1f}%"])
                ws.append(["Attendance Rate", f"{m.attendance_rate:.1f}%"])
                ws.append(["Trend", m.trend_direction.capitalize()])
                ws.append([])

            # Lesson records table
            headers = ["Date", "Attendance", "Grade", "Feedback"]
            ws.append(headers)
            header_row = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

            for r in cd["records"]:
                att = r.attendance_status.value if hasattr(r.attendance_status, "value") else r.attendance_status
                ws.append([
                    r.lesson_date.strftime("%Y-%m-%d"),
                    att,
                    round(r.grade_value, 1) if r.grade_value is not None else "",
                    r.teacher_feedback or "",
                ])

            # AI Alerts section
            if cd["alerts"]:
                ws.append([])
                ws.append(["AI Alerts"])
                alert_headers = ["Level", "Message", "Recommendation", "Date"]
                ws.append(alert_headers)
                hr2 = ws.max_row
                for col in range(1, len(alert_headers) + 1):
                    cell = ws.cell(row=hr2, column=col)
                    cell.fill = PatternFill("solid", fgColor="F97316")
                    cell.font = HEADER_FONT
                for a in cd["alerts"]:
                    lvl = a.alert_level.value if hasattr(a.alert_level, "value") else a.alert_level
                    ws.append([lvl, a.message, a.recommendation, a.created_at.strftime("%Y-%m-%d")])

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        # Remove default empty sheet if we created at least one
        if course_data:
            wb.remove(default_sheet)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"report_{student_id_from(student)}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fname

    except ImportError:
        logger.warning("openpyxl not installed — falling back to CSV")
        return _build_csv(student, course_data)


# ── Course performance summary ─────────────────────────────────────────────────

async def generate_course_report(course_id: str, fmt: str = "csv") -> tuple[bytes, str, str]:
    """Aggregate performance summary for an entire course."""
    from app.models import Course, Enrollment, EnrollmentStatusEnum
    from app.repositories import ProgressMetricsRepository, UserRepository

    course = await Course.get(course_id)
    if not course:
        raise ValueError("Course not found")

    enrollments = await Enrollment.find(
        Enrollment.course_id == course_id,
        Enrollment.status == EnrollmentStatusEnum.ACTIVE,
    ).to_list()

    rows = []
    for e in enrollments:
        student = await UserRepository.get_by_id(e.student_id)
        metrics = await ProgressMetricsRepository.get(e.student_id, course_id)
        rows.append({
            "name": student.full_name() if student else e.student_id,
            "email": student.email if student else "",
            "avg_grade": metrics.average_grade if metrics else 0.0,
            "attendance": metrics.attendance_rate if metrics else 0.0,
            "trend": metrics.trend_direction if metrics else "stable",
        })

    if fmt == "xlsx":
        return _build_course_xlsx(course, rows)
    return _build_course_csv(course, rows)


def _build_course_csv(course, rows: list) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["IQ PLUS — Course Performance Report"])
    w.writerow([f"Course: {course.name} ({course.code})"])
    w.writerow([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
    w.writerow([])
    w.writerow(["Student", "Email", "Avg Grade", "Attendance %", "Trend"])
    for r in rows:
        w.writerow([r["name"], r["email"], f"{r['avg_grade']:.1f}", f"{r['attendance']:.1f}", r["trend"]])
    fname = f"course_{str(course.id)[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return buf.getvalue().encode("utf-8"), "text/csv", fname


def _build_course_xlsx(course, rows: list) -> tuple[bytes, str, str]:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = course.code[:31]

        ws.append([f"IQ PLUS — Course Performance Report"])
        ws.append([f"Course: {course.name} ({course.code})"])
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
        ws.append([])
        headers = ["Student", "Email", "Avg Grade", "Attendance %", "Trend"]
        ws.append(headers)
        hr = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=hr, column=col)
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.font = Font(color="FFFFFF", bold=True)

        for r in rows:
            ws.append([r["name"], r["email"], round(r["avg_grade"], 1), round(r["attendance"], 1), r["trend"]])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"course_{str(course.id)[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fname
    except ImportError:
        return _build_course_csv(course, rows)


# ── Attendance report ─────────────────────────────────────────────────────────

async def generate_attendance_report(course_id: str, student_id: str | None = None, fmt: str = "csv") -> tuple[bytes, str, str]:
    """Per-course (optionally per-student) attendance report."""
    from app.models import Course, LessonRecord
    from app.repositories import LessonRecordRepository, UserRepository

    course = await Course.get(course_id)
    if not course:
        raise ValueError("Course not found")

    if student_id:
        records = await LessonRecordRepository.get_by_student_course(student_id, course_id)
        student = await UserRepository.get_by_id(student_id)
        grouped = {student.full_name() if student else student_id: records}
    else:
        raw = await LessonRecord.find(LessonRecord.course_id == course_id).sort(-LessonRecord.lesson_date).to_list()
        grouped: dict = {}
        for r in raw:
            grouped.setdefault(r.student_id, []).append(r)

    if fmt == "xlsx":
        return _build_attendance_xlsx(course, grouped)
    return _build_attendance_csv(course, grouped)


def _build_attendance_csv(course, grouped: dict) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["IQ PLUS — Attendance Report"])
    w.writerow([f"Course: {course.name} ({course.code})"])
    w.writerow([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
    w.writerow([])
    w.writerow(["Student", "Date", "Status"])
    for label, records in grouped.items():
        for r in records:
            att = r.attendance_status.value if hasattr(r.attendance_status, "value") else r.attendance_status
            w.writerow([label, r.lesson_date.strftime("%Y-%m-%d"), att])
        w.writerow([])
    fname = f"attendance_{str(course.id)[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
    return buf.getvalue().encode("utf-8"), "text/csv", fname


def _build_attendance_xlsx(course, grouped: dict) -> tuple[bytes, str, str]:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        ws.append([f"IQ PLUS — Attendance Report"])
        ws.append([f"Course: {course.name} ({course.code})"])
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
        ws.append([])
        headers = ["Student", "Date", "Status"]
        ws.append(headers)
        hr = ws.max_row
        for col in range(1, 4):
            cell = ws.cell(row=hr, column=col)
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.font = Font(color="FFFFFF", bold=True)

        for label, records in grouped.items():
            for r in records:
                att = r.attendance_status.value if hasattr(r.attendance_status, "value") else r.attendance_status
                ws.append([label, r.lesson_date.strftime("%Y-%m-%d"), att])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"attendance_{str(course.id)[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fname
    except ImportError:
        return _build_attendance_csv(course, grouped)
